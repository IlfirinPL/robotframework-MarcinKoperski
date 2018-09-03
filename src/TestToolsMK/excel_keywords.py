#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Copyright (c) 2015 Cutting Edge QA Marcin Koperski
import io
import os
import time

import openpyxl
import unicodecsv as csv
from openpyxl import load_workbook
from robot.api import logger
from robot.libraries import DateTime
from robot.utils import asserts

from TestToolsMK.robot_instances import *


class ExcelKeywords(object):

    def __init__(self):
        self.wb = None  # type: openpyxl.workbook
        self.tb = None
        self.sheetNum = None
        self.sheetNames = None
        self.currentSheet = None  # type: openpyxl.worksheet
        self.fileName = None
        self.mode = None
        if os.name is "nt":
            self.tmpDir = "Temp"
        else:
            self.tmpDir = "tmp"

    def open_excel(self, filename, read_only=True, **kwargs):
        """
        Open spread excel and return spreadsheet names.
        Select first spreadsheet as current
        :param self: 
        :param filename: 
        :param kwargs:
        :return: 
        @param filename:
        @param read_only:
        """
        self.mode = read_only
        self.fileName = filename
        self.wb = load_workbook(self.fileName, read_only, **kwargs)
        self.sheetNames = self.wb.get_sheet_names()
        self.select_spreadsheet("")
        return self.sheetNames

    def save_working_excel(self, filename):
        """
        Save working file"
        """
        self.fileName = filename
        self.wb.save(self.fileName)
        self.open_excel(self.fileName, self.mode)

    def select_spreadsheet(self, name=""):

        if name == "":
            name = self.sheetNames[0]
            self.currentSheet = self.wb[name]
        else:
            self.currentSheet = self.wb[name]

    def get_all_values(self):

        table = []
        for row in self.currentSheet.rows:

            single_row = []
            for cell in row:
                single_row.append(cell.value)

            table.append(single_row)
        return table

    def get_cell_data_by_coordinates(self, column, row):
        """
        will use currently selected spreadsheet to change use method
        | select SpreadSheet | mySpreadSheet |
        | ${value}| get_cell_data_by_coordinates | A | 1 |

        """
        cell_value = self.currentSheet[column + row].value
        return cell_value

    def edit_data_by_coordinates(self, col_num, row_num, value=""):
        """
        Edit cell in open excel file by coordinates column and row number. Edit will use current spreadsheet
        | select SpreadSheet | mySpreadSheet |
        | ${old value}| edit_data_by_coordinates | 1 | 1 | Example |
        return old value
        """
        old = self.currentSheet.cell(row=int(row_num), column=int(col_num)).value
        self.currentSheet.cell(row=int(row_num), column=int(col_num)).value = value
        return old


